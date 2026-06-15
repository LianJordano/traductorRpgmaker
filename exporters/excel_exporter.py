"""Export extraction results to Excel (.xlsx) with formatting."""
from __future__ import annotations
from pathlib import Path

from core.models import ExtractionResult
from exporters.base import BaseExporter
from core import logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

HEADERS = ["UID", "File", "Context", "Original", "Translation", "Status"]
COL_WIDTHS = [40, 20, 35, 60, 60, 12]

HEADER_FILL = "1F3864"
HEADER_FONT_COLOR = "FFFFFF"
PENDING_FILL = "FFF2CC"
TRANSLATED_FILL = "E2EFDA"
ERROR_FILL = "FCE4D6"


class ExcelExporter(BaseExporter):
    def export(self, result: ExtractionResult, output_path: Path) -> Path:
        if not _OPENPYXL_OK:
            logger.warning("openpyxl not installed; falling back to CSV export")
            from exporters.csv_exporter import CsvExporter
            # Force .csv extension so the file is not named .xlsx with CSV content inside
            csv_path = Path(output_path).with_suffix(".csv")
            return CsvExporter().export(result, csv_path)

        output_path = Path(output_path)
        if output_path.is_dir():
            output_path = output_path / "translation.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translations"

        # Header row
        header_font = Font(bold=True, color=HEADER_FONT_COLOR)
        header_fill = PatternFill("solid", fgColor=HEADER_FILL)
        for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = w

        ws.freeze_panes = "A2"

        status_fills = {
            "pending": PatternFill("solid", fgColor=PENDING_FILL),
            "translated": PatternFill("solid", fgColor=TRANSLATED_FILL),
            "error": PatternFill("solid", fgColor=ERROR_FILL),
        }
        wrap = Alignment(wrap_text=True, vertical="top")

        for row_idx, entry in enumerate(result.entries, 2):
            values = [
                entry.uid, entry.file, entry.context,
                entry.original, entry.translation, entry.status,
            ]
            fill = status_fills.get(entry.status, PatternFill())
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = wrap
                if fill:
                    cell.fill = fill

        ws.auto_filter.ref = ws.dimensions
        wb.save(output_path)
        return output_path
