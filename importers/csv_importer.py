"""Import translations from a CSV/JSON/Excel file back into an ExtractionResult."""
from __future__ import annotations
import csv
from pathlib import Path

from core.models import ExtractionResult, TextEntry
from core import logger

# Only line endings are trimmed from an imported translation: leading and
# trailing spaces are what RPG Maker uses to align menu entries, so stripping
# them visibly shifts the game's UI.
_EOL = "\r\n"


def _clean(value) -> str:
    return str(value or "").strip(_EOL)


def import_csv(csv_path: Path, result: ExtractionResult) -> int:
    """Merge translations from CSV into result. Returns count of updated entries."""
    uid_map: dict[str, TextEntry] = {e.uid: e for e in result.entries}
    updated = 0
    unknown = 0

    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            uid = (row.get("uid") or "").strip()
            translation = _clean(row.get("translation"))
            if not translation.strip():
                continue
            entry = uid_map.get(uid)
            if entry is None:
                unknown += 1
                continue
            entry.translation = translation
            entry.status = "translated"
            updated += 1

    if unknown:
        logger.warning(
            f"{unknown} rows in the CSV do not match any extracted text "
            f"(the game files may have changed since the export)."
        )
    logger.success(f"Imported {updated} translations from CSV")
    return updated


def import_json(json_path: Path, result: ExtractionResult) -> int:
    """Import from a JSON file (same format as our JSON export)."""
    import json
    uid_map: dict[str, TextEntry] = {e.uid: e for e in result.entries}
    updated = 0

    data = json.loads(json_path.read_text(encoding="utf-8"))
    for entry_data in data.get("entries", []):
        uid = entry_data.get("uid", "")
        translation = _clean(entry_data.get("translation"))
        if uid in uid_map and translation.strip():
            uid_map[uid].translation = translation
            uid_map[uid].status = "translated"
            updated += 1

    logger.success(f"Imported {updated} translations from JSON")
    return updated


def import_excel(xlsx_path: Path, result: ExtractionResult) -> int:
    """Import from an Excel file."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed for Excel import")
        return 0

    uid_map: dict[str, TextEntry] = {e.uid: e for e in result.entries}
    updated = 0

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        uid_col = headers.index("UID")
        trans_col = headers.index("Translation")
    except ValueError:
        logger.error("Excel file missing UID or Translation column")
        return 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        uid = str(row[uid_col] or "").strip()
        translation = _clean(row[trans_col])
        if uid in uid_map and translation.strip():
            uid_map[uid].translation = translation
            uid_map[uid].status = "translated"
            updated += 1

    logger.success(f"Imported {updated} translations from Excel")
    return updated
